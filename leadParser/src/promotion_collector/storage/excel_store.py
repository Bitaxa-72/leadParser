from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from promotion_collector.models import BusinessRecord


HEADERS = [
    "Название бизнеса",
    "Тип бизнеса",
    "Город",
    "Сайт",
    "Почта",
    "Telegram",
    "VK",
    "Instagram",
    "Facebook",
    "WhatsApp",
    "YouTube",
    "Другая соцсеть",
    "Телефон",
    "Адрес",
    "Источник данных",
    "Название источника",
    "Дата сбора",
]


class ExcelStore:
    def __init__(self, path: Path) -> None:
        self.path = path

    def append_iteration_sheet(self, records: list[BusinessRecord]) -> str | None:
        if not records:
            return None

        self.path.parent.mkdir(parents=True, exist_ok=True)
        workbook = load_workbook(self.path) if self.path.exists() else Workbook()
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
            del workbook["Sheet"]

        sheet_name = _next_sheet_name(workbook.sheetnames)
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(HEADERS)
        for record in records:
            worksheet.append(record.to_excel_row())

        _format_sheet(worksheet)
        workbook.save(self.path)
        return sheet_name


def _next_sheet_name(existing_names: list[str]) -> str:
    base = datetime.now().strftime("%Y-%m-%d")
    if base not in existing_names:
        return base
    index = 2
    while True:
        candidate = f"{base}_{index}"
        if candidate not in existing_names:
            return candidate
        index += 1


def _format_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = {
        1: 30,
        2: 22,
        3: 16,
        4: 34,
        5: 26,
        6: 28,
        7: 28,
        8: 28,
        9: 28,
        10: 28,
        11: 28,
        12: 34,
        13: 20,
        14: 34,
        15: 42,
        16: 24,
        17: 24,
    }
    for index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(index)].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
